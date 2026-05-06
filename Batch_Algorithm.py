# -*- coding: utf-8 -*-
"""
간호학과 78명( A1–A37, B1–B41 )의 세 과목-병원-기간 배치를
모든 제약조건(학생별 3과목 1회, 기간 중복 금지, 병원 중복 금지,
셀 최소 2명, 2학기 셀 의무 사용, Max/Min 용량) 을 만족하도록
정수계획으로 최적화한 뒤 CSV 출력
"""
import csv
import json
from collections import defaultdict
from itertools import product
from pathlib import Path

from pulp import (
    LpProblem, LpVariable, LpBinary, lpSum,
    LpStatusOptimal, PULP_CBC_CMD
)

# -------------------- 1. 기본 데이터 --------------------
# web_config.json 파일에서 설정을 불러옵니다.
p = Path("web_config.json")
if not p.exists():
    raise FileNotFoundError("web_config.json 파일이 존재하지 않습니다. 웹 페이지에서 설정을 저장한 후 다시 시도하세요.")

try:
    with p.open("r", encoding="utf-8") as f:
        cfg = json.load(f)
except Exception as e:
    raise RuntimeError(f"web_config.json 파일을 읽는 중 오류가 발생했습니다: {e}")

if not cfg:
    raise RuntimeError("web_config.json에서 필요한 설정을 읽을 수 없습니다.")

num_periods = int(cfg.get("numPeriods") or cfg.get("oneA", {}).get("numPeriods") or 0)
if num_periods == 0:
    raise RuntimeError("web_config.json에서 기간 수(numPeriods)를 읽을 수 없습니다.")

periods = [f"T{i}" for i in range(1, num_periods + 1)]

assignments = cfg.get("assignments", [])
def get_semester(i: int) -> str:
    for a in assignments:
        if int(a.get("period", 0)) == i:
            return str(a.get("semester") or "")
    return ""
second_semester = {f"T{i}" for i in range(1, num_periods + 1) if get_semester(i) == "2"}

numA = int(cfg.get("numA", 0))
numB = int(cfg.get("numB", 0))
if numA == 0 and numB == 0:
    raise RuntimeError("web_config.json에서 학생 수(numA, numB)를 읽을 수 없습니다.")

students_A = [f"A{i:02d}" for i in range(1, numA + 1)]
students_B = [f"B{i:02d}" for i in range(1, numB + 1)]
students = students_A + students_B

def rule_for_period(i: int):
    cls = ""
    for a in assignments:
        if int(a.get("period", 0)) == i:
            cls = (a.get("class") or "").strip()
            break
    if cls == "A":
        return lambda s: s.startswith("A")
    if cls == "B":
        return lambda s: s.startswith("B")
    return lambda s: True  # A&B 또는 미지정 → 모두 가능

period_ok = {f"T{i}": rule_for_period(i) for i in range(1, num_periods + 1)}

oneA = cfg.get("oneA", {})
max_rows = oneA.get("max", [])
min_rows = oneA.get("min", [])

if not max_rows:
    raise RuntimeError("web_config.json에서 최대 용량 데이터(max)를 읽을 수 없습니다.")

# 3학년 1학기, 3학년 2학기, 4학년 1학기, 4학년 2학기 여부 확인
selected_semester = cfg.get("selectedSemester", {})
selected_grade = int(selected_semester.get("grade", 0))
selected_semester_num = int(selected_semester.get("semester", 0))
is_grade3_semester1 = (selected_grade == 3 and selected_semester_num == 1)
is_grade3_semester2 = (selected_grade == 3 and selected_semester_num == 2)
is_grade4_semester1 = (selected_grade == 4 and selected_semester_num == 1)
is_grade4_semester2 = (selected_grade == 4 and selected_semester_num == 2)
# 각 행별 최대 인원 체크가 필요한 학기들
use_row_based_capacity = is_grade3_semester1 or is_grade3_semester2 or is_grade4_semester1 or is_grade4_semester2
# 특정 학기 그룹 (3학년 2학기, 4학년 1학기, 4학년 2학기)
special_semesters = is_grade3_semester2 or is_grade4_semester1 or is_grade4_semester2

# 중복 배치 허용 병원 목록 추출
allow_duplicate_hospitals = {
    row.get("hospital", "").strip()
    for row in max_rows
    if row.get("allowDuplicate", False) and row.get("hospital", "").strip()
}

def rows_to_csv(rows: list) -> str:
    """행 데이터를 CSV 형식 문자열로 변환"""
    lines = []
    for r in rows:
        if is_grade3_semester1:
            # 3학년 1학기: hospital(병원명)을 course 필드에, department(부서명)을 hospital 필드에 저장
            course = r.get("hospital", "")
            hosp = r.get("department", "")
        elif special_semesters:
            # 3학년 2학기, 4학년 1학기, 4학년 2학기: subject(과목명)과 hospital(병원명), department는 무시
            course = r.get("subject", "")
            hosp = r.get("hospital", "")
        else:
            # 일반 모드: subject(과목명)과 hospital(병원명)
            course = r.get("subject", "")
            hosp = r.get("hospital", "")
        values = r.get("values", [])
        # pad/truncate to num_periods
        values = (values + [0] * num_periods)[:num_periods]
        line = ",".join([course, hosp] + [str(int(v)) for v in values])
        lines.append(line)
    return "\n".join(lines)

RAW_MAX = rows_to_csv(max_rows)
RAW_MIN = rows_to_csv(min_rows) if min_rows else RAW_MAX

def parse_rows(raw: str, max_rows_data: list = None):
    """CSV 형식 문자열을 행 데이터 리스트로 파싱"""
    rows = []
    for idx, line in enumerate(raw.splitlines()):
        course, hosp, *nums = line.strip().split(",")
        row_data = {
            "id": f"R{idx:02d}",
            "course": course,
            "hospital": hosp,
            "capacity": {p: int(x) for p, x in zip(periods, nums)}
        }
        # 3학년 1학기인 경우 deptType 정보 추가
        if max_rows_data and idx < len(max_rows_data):
            dept_type = max_rows_data[idx].get("deptType", "")
            if dept_type:
                row_data["deptType"] = dept_type
        rows.append(row_data)
    return rows

rows_max = parse_rows(RAW_MAX, max_rows if is_grade3_semester1 else None)
if RAW_MIN.strip():
    rows_min = parse_rows(RAW_MIN)
else:
    # min_rows가 비어있으면 rows_max와 같은 구조로 0으로 채워진 용량 생성
    rows_min = [
        {
            "id": r_max["id"],
            "course": r_max["course"],
            "hospital": r_max["hospital"],
            "capacity": {p: 0 for p in periods}
        }
        for r_max in rows_max
    ]

# ----- 1-B. 셀별 최저 인원 계산(규칙 4·5 반영) -----
for r_max, r_min in zip(rows_max, rows_min):
    for p in periods:
        rule5 = 2 if (p in second_semester and r_max["capacity"][p] > 0) else 0
        r_min["capacity"][p] = max(r_min["capacity"][p], rule5)

# ---------------- 2. MILP 모델 ----------------------
model = LpProblem("Nursing-Student-Scheduling")

# 변수: assign[s, row_id, period]  ∈ {0,1}
assign = {}
for s, r in product(students, rows_max):
    for p in periods:
        if r["capacity"][p] == 0:
            continue
        if not period_ok[p](s):
            continue
        key = (s, r["id"], p)
        assign[key] = LpVariable(f"x_{s}_{r['id']}_{p}", 0, 1, LpBinary)

# 변수: y_(row,period) – 셀 사용 여부 (0: 미사용, 1: 사용)
y = {}
for r in rows_max:
    for p in periods:
        if r["capacity"][p] == 0:
            continue
        y[(r['id'], p)] = LpVariable(f"y_{r['id']}_{p}", 0, 1, LpBinary)

# 성능 최적화: row_id -> row_data 매핑 생성
row_id_to_index = {r['id']: idx for idx, r in enumerate(rows_max)}
row_id_to_row = {r['id']: r for r in rows_max}

# 2-A. 학생 배치 규칙
if is_grade3_semester1:
    # 3학년 1학기: 학생은 정확히 2번 배치되어야 함
    for s in students:
        all_assigns_for_student = [assign[k] for k in assign if k[0] == s]
        if all_assigns_for_student:
            model += lpSum(all_assigns_for_student) == 2, f"TotalAssignments_{s}"
else:
    # 일반 모드 및 특정 학기: 학생·과목별 1회 수강 (과목 중복 금지)
    all_courses = sorted(set(r['course'] for r in rows_max))
    for s in students:
        for course in all_courses:
            model += (
                lpSum(assign[k] for k in assign
                      if k[0] == s and row_id_to_row[k[1]]['course'] == course) == 1
            )

# 2-B. 학생·기간 중복 금지
for s in students:
    for p in periods:
        model += lpSum(assign[k] for k in assign if k[0] == s and k[2] == p) <= 1

# 2-C. 학생·병원(기관) 중복 금지 (중복 배치 허용 병원 제외)
if is_grade3_semester1:
    # 3학년 1학기: course 필드에 실제 병원명이 저장됨
    hospitals = {r['course'] for r in rows_max}
    for s in students:
        for hosp in hospitals:
            if hosp not in allow_duplicate_hospitals:
                model += lpSum(assign[k] for k in assign
                               if k[0] == s and row_id_to_row[k[1]]['course'] == hosp) <= 1
else:
    # 일반 모드: hospital 필드에 병원명이 저장됨
    hospitals = {r['hospital'] for r in rows_max}
    for s in students:
        for hosp in hospitals:
            if hosp not in allow_duplicate_hospitals:
                model += lpSum(assign[k] for k in assign
                               if k[0] == s and row_id_to_row[k[1]]['hospital'] == hosp) <= 1

# 2-D. 셀 용량(최대·최소 및 '0 또는 ≥2')
if use_row_based_capacity:
    # 3학년 1학기, 3학년 2학기, 4학년 1학기, 4학년 2학기: 각 행(부서)별로 최대 인원을 독립적으로 체크
    for r_max in rows_max:
        rid = r_max['id']
        for p in periods:
            cap_max = r_max["capacity"][p]
            if cap_max == 0:
                continue
            cell_vars = [assign[k] for k in assign if k[1] == rid and k[2] == p]
            if cell_vars:
                model += lpSum(cell_vars) <= cap_max, f"MaxCapacity_Row_{rid}_Period_{p}"
                # 3학년 1학기: 모든 배치 자리에 최소 2명 배치
                if is_grade3_semester1:
                    model += lpSum(cell_vars) >= 2, f"MinCapacity_Row_{rid}_Period_{p}"
else:
    # 일반 모드: 기존 로직 (최소 인원 및 0 또는 ≥2 제약 포함)
    for r_max, r_min in zip(rows_max, rows_min):
        rid = r_max['id']
        for p in periods:
            cap_max = r_max["capacity"][p]
            if cap_max == 0:
                continue
            cell_vars = [assign[k] for k in assign if k[1] == rid and k[2] == p]
            model += lpSum(cell_vars) <= cap_max
            cap_min = r_min["capacity"][p]
            model += lpSum(cell_vars) >= cap_min
            model += lpSum(cell_vars) >= 2 * y[(rid, p)]
            model += lpSum(cell_vars) <= cap_max * y[(rid, p)]

# 2-E. 목적함수: 우선순위 가중치 (3학년 1학기만)
if is_grade3_semester1:
    # 각 학생별 배치 패턴에 따른 페널티 계산
    student_penalties = []
    
    for s in students:
        # 각 학생의 내과/외과/내과&외과 배치 횟수 계산
        internal_count_list = []
        external_count_list = []
        both_count_list = []
        
        for r in rows_max:
            dept_type = r.get("deptType", "")
            rid = r['id']
            for p in periods:
                key = (s, rid, p)
                if key in assign:
                    if dept_type == "internal":
                        internal_count_list.append(assign[key])
                    elif dept_type == "external":
                        external_count_list.append(assign[key])
                    elif dept_type == "both":
                        both_count_list.append(assign[key])
        
        internal_count = lpSum(internal_count_list) if internal_count_list else 0
        external_count = lpSum(external_count_list) if external_count_list else 0
        both_count = lpSum(both_count_list) if both_count_list else 0
        
        # 우선순위 패턴 판별을 위한 보조 변수
        pattern1 = LpVariable(f"pattern1_{s}", cat='Binary')
        pattern2 = LpVariable(f"pattern2_{s}", cat='Binary')
        pattern3 = LpVariable(f"pattern3_{s}", cat='Binary')
        pattern4 = LpVariable(f"pattern4_{s}", cat='Binary')
        
        # 패턴 1: internal=1 and external=1 and both=0
        model += internal_count <= 1 + 10 * (1 - pattern1)
        model += internal_count >= 1 - 10 * (1 - pattern1)
        model += external_count <= 1 + 10 * (1 - pattern1)
        model += external_count >= 1 - 10 * (1 - pattern1)
        model += both_count <= 10 * (1 - pattern1)
        
        # 패턴 2: (internal=1 and both=1) or (external=1 and both=1)
        pattern2a = LpVariable(f"pattern2a_{s}", cat='Binary')
        pattern2b = LpVariable(f"pattern2b_{s}", cat='Binary')
        
        model += internal_count <= 1 + 10 * (1 - pattern2a)
        model += internal_count >= 1 - 10 * (1 - pattern2a)
        model += both_count <= 1 + 10 * (1 - pattern2a)
        model += both_count >= 1 - 10 * (1 - pattern2a)
        
        model += external_count <= 1 + 10 * (1 - pattern2b)
        model += external_count >= 1 - 10 * (1 - pattern2b)
        model += both_count <= 1 + 10 * (1 - pattern2b)
        model += both_count >= 1 - 10 * (1 - pattern2b)
        
        model += pattern2 <= pattern2a + pattern2b
        model += pattern2 >= pattern2a
        model += pattern2 >= pattern2b
        
        # 패턴 3: both=2
        model += both_count <= 2 + 10 * (1 - pattern3)
        model += both_count >= 2 - 10 * (1 - pattern3)
        
        # 패턴 4: internal=2 or external=2
        pattern4a = LpVariable(f"pattern4a_{s}", cat='Binary')
        pattern4b = LpVariable(f"pattern4b_{s}", cat='Binary')
        
        model += internal_count <= 2 + 10 * (1 - pattern4a)
        model += internal_count >= 2 - 10 * (1 - pattern4a)
        model += external_count <= 2 + 10 * (1 - pattern4b)
        model += external_count >= 2 - 10 * (1 - pattern4b)
        
        model += pattern4 <= pattern4a + pattern4b
        model += pattern4 >= pattern4a
        model += pattern4 >= pattern4b
        
        # 패턴들이 상호 배타적이어야 함
        model += pattern1 + pattern2 + pattern3 + pattern4 == 1
        
        # 페널티 계산: 패턴 1=0, 패턴 2=10, 패턴 3=70, 패턴 4=95
        penalty = 0 * pattern1 + 10 * pattern2 + 70 * pattern3 + 95 * pattern4
        student_penalties.append(penalty)
    
    if student_penalties:
        model += lpSum(student_penalties)

# ---------------- 3. 풀기 ---------------------------
solver = PULP_CBC_CMD(msg=1, timeLimit=120)
status = model.solve(solver)

if status != LpStatusOptimal:
    raise RuntimeError("배치를 찾지 못했습니다. 제약을 완화하거나 용량을 재검토하세요.")

# ---------------- 4. 결과 CSV 작성 -------------------
header = ["과목명", "병원명"] + periods
csv_rows = []

# 행별-기간별 학생 이름 모으기
who_in = defaultdict(lambda: defaultdict(list))

for (s, rid, p), var in assign.items():
    if var.value() == 1:
        who_in[rid][p].append(s)

for r in rows_max:
    rid = r['id']
    line = [r['course'], r['hospital']]
    for p in periods:
        if r['capacity'][p] == 0:
            line.append("")
        else:
            stu_list = " ".join(sorted(who_in[rid][p]))
            line.append(stu_list)
    csv_rows.append(line)

out_path = Path("배치결과.csv")
with out_path.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(csv_rows)

# ---------------- 5. Excel 파일 생성 (배치결과, 명단, 최종결과) ----------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl 라이브러리가 필요합니다. 'pip install openpyxl' 명령으로 설치해주세요.")

final_excel_path = Path("배치결과_최종.xlsx")
wb = Workbook()

# 첫 번째 시트: 배치결과 (학생 ID)
ws1 = wb.active
ws1.title = "배치결과"

# 헤더 작성
ws1.append(header)
header_style = {
    'font': Font(bold=True),
    'alignment': Alignment(horizontal='center', vertical='center'),
    'fill': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
}
for cell in ws1[1]:
    for attr, value in header_style.items():
        setattr(cell, attr, value)

# 데이터 작성
for row_data in csv_rows:
    ws1.append(row_data)

# 열 너비 자동 조정
for col in range(1, len(header) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 15

# 두 번째 시트: 명단 입력용
ws2 = wb.create_sheet("명단")

# A반 헤더
ws2.append(["A반"])
ws2.append(["번호", "학번", "이름"])
a_header_style = {
    'font': Font(bold=True),
    'alignment': Alignment(horizontal='center', vertical='center'),
    'fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
}
for cell in ws2[2]:
    for attr, value in a_header_style.items():
        setattr(cell, attr, value)

# A반 빈 행 추가
for i in range(1, numA + 1):
    ws2.append([f"A{i:02d}", "", ""])

# B반 헤더
ws2.append([])
ws2.append(["B반"])
ws2.append(["번호", "학번", "이름"])
b_header_row = len(ws2['A'])
b_header_style = {
    'font': Font(bold=True),
    'alignment': Alignment(horizontal='center', vertical='center'),
    'fill': PatternFill(start_color="FFE6F3", end_color="FFE6F3", fill_type="solid")
}
for cell in ws2[b_header_row]:
    for attr, value in b_header_style.items():
        setattr(cell, attr, value)

# B반 빈 행 추가
for i in range(1, numB + 1):
    ws2.append([f"B{i:02d}", "", ""])

# 열 너비 조정
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15

# 세 번째 시트: 이름으로 치환된 배치결과
ws3 = wb.create_sheet("배치결과_이름")

# 헤더 작성
ws3.append(header)
for cell in ws3[1]:
    for attr, value in header_style.items():
        setattr(cell, attr, value)

# 배치결과를 세 번째 시트에 작성 (Excel 수식 사용)
def normalize_student_id(tok: str) -> str:
    """학생 ID를 A01 형식으로 정규화"""
    tok_clean = tok.strip()
    if len(tok_clean) >= 2:
        prefix = tok_clean[0]
        try:
            num = int(tok_clean[1:])
            return f"{prefix}{num:02d}"
        except ValueError:
            return tok_clean
    return tok_clean

for row_idx, row_data in enumerate(csv_rows, start=2):
    ws3.cell(row=row_idx, column=1, value=row_data[0])
    ws3.cell(row=row_idx, column=2, value=row_data[1])
    
    for col_idx in range(3, len(row_data) + 1):
        cell_value = str(row_data[col_idx - 1]).strip() if col_idx - 1 < len(row_data) else ""
        
        if cell_value:
            tokens = cell_value.split()
            formula_parts = [
                f'IFERROR(VLOOKUP("{normalize_student_id(tok)}",명단!A:C,3,FALSE),"{normalize_student_id(tok)}")'
                for tok in tokens if tok.strip()
            ]
            
            if formula_parts:
                if len(formula_parts) == 1:
                    ws3.cell(row=row_idx, column=col_idx).value = f'={formula_parts[0]}'
                else:
                    formula_str = "=" + '&" "&'.join(formula_parts)
                    ws3.cell(row=row_idx, column=col_idx).value = formula_str
        else:
            ws3.cell(row=row_idx, column=col_idx, value="")

# 세 번째 시트 열 너비 조정
for col in range(1, len(header) + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# 네 번째 시트: 학생별 배치 요약 (3학년 1학기, 3학년 2학기, 4학년 1학기, 4학년 2학기)
if use_row_based_capacity:
    ws4 = wb.create_sheet("학생별배치요약")
    
    # 첫 번째 행: 전체 기간 정보
    period_row = ["기간"] + list(periods)
    ws4.append(period_row)
    period_row_style = {
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    }
    for cell in ws4[1]:
        for attr, value in period_row_style.items():
            setattr(cell, attr, value)
    
    # 헤더 작성 (두 번째 행)
    header_row = ["학생 ID"] + list(periods)
    ws4.append(header_row)
    for cell in ws4[2]:
        for attr, value in header_style.items():
            setattr(cell, attr, value)
    
    # 각 학생별 배치 정보 수집 (기간별로)
    student_period_assignments = defaultdict(lambda: defaultdict(list))
    
    for (s, rid, p), var in assign.items():
        if var.value() == 1:
            row_data = row_id_to_row.get(rid)
            if row_data:
                if is_grade3_semester1:
                    institution = row_data.get('course', '')
                    department = row_data.get('hospital', '')
                    dept_type = row_data.get('deptType', '')
                elif special_semesters:
                    institution = row_data.get('hospital', '')
                    # 원본 max_rows에서 부서 정보 가져오기
                    department = ""
                    dept_type = ""
                    try:
                        row_idx = row_id_to_index.get(rid, -1)
                        if 0 <= row_idx < len(max_rows):
                            department = max_rows[row_idx].get('department', '') or ""
                            dept_type = max_rows[row_idx].get('deptType', '') or ""
                    except (ValueError, IndexError):
                        department = ""
                        dept_type = ""
                else:
                    institution = row_data.get('hospital', '')
                    department = ""
                    dept_type = ""
                student_period_assignments[s][p].append((institution, department, dept_type))
    
    # 부서 타입을 한국어로 변환하는 함수
    def dept_type_to_korean(dept_type):
        if dept_type == "internal":
            return "내과"
        elif dept_type == "external":
            return "외과"
        elif dept_type == "both":
            return "내과&외과"
        return ""
    
    # 학생별로 정렬하고 배치 정보 작성
    for student in sorted(students):
        # 3학년 1학기: 전체 배치의 부서 타입 정보 집계
        if is_grade3_semester1:
            total_dept_type_counts = defaultdict(int)
            for p in periods:
                assignments = student_period_assignments[student].get(p, [])
                for inst, dept, dept_type in assignments:
                    if dept_type:
                        korean_type = dept_type_to_korean(dept_type)
                        if korean_type:
                            total_dept_type_counts[korean_type] += 1
            
            # 학생 ID 열에 부서 타입 정보 추가
            dept_type_parts = []
            for dept_type_kr in ["내과", "외과", "내과&외과"]:
                count = total_dept_type_counts.get(dept_type_kr, 0)
                if count > 0:
                    dept_type_parts.append(f"{dept_type_kr}{count}")
            
            if dept_type_parts:
                student_id_with_type = f"{student}({', '.join(dept_type_parts)})"
            else:
                student_id_with_type = student
        else:
            student_id_with_type = student
        
        row_data = [student_id_with_type]
        for p in periods:
            assignments = student_period_assignments[student].get(p, [])
            if assignments:
                # 기존처럼 병원과 부서명만 표시
                assignment_strs = [
                    f"{inst}({dept})" if inst and dept else inst
                    for inst, dept, dt in assignments
                    if inst
                ]
                row_data.append(" / ".join(assignment_strs))
            else:
                row_data.append("")
        ws4.append(row_data)
    
    # 열 너비 조정
    ws4.column_dimensions['A'].width = 12
    for col_idx in range(2, ws4.max_column + 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 40

# 첫 번째 시트(배치결과)에서 최대 인원수에 충족하는 셀에 연한 노란색 채우기 (3학년 1학기만)
if is_grade3_semester1:
    light_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    ws1 = wb["배치결과"]
    
    # 헤더 행에서 기간 정보 추출
    period_columns = {
        col_idx: ws1.cell(row=1, column=col_idx).value
        for col_idx in range(3, ws1.max_column + 1)
        if ws1.cell(row=1, column=col_idx).value
    }
    
    # 각 행의 배치 정보 확인
    for row_idx in range(2, ws1.max_row + 1):
        hospital = ws1.cell(row=row_idx, column=2).value or ""
        
        # 해당 병원/부서의 rows_max 데이터 찾기
        for r in rows_max:
            if r.get('course') == hospital or r.get('hospital') == hospital:
                for col_idx, period_name in period_columns.items():
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value or ""
                    
                    max_capacity = r.get("capacity", {}).get(period_name, 0)
                    if max_capacity > 0:
                        assigned_count = len([x for x in cell_value.split() if x.strip()]) if cell_value else 0
                        if assigned_count == max_capacity:
                            cell.fill = light_yellow
                break

# Excel 파일 저장
wb.save(final_excel_path)
print(f"완료: {final_excel_path.resolve()}")
print("  - 첫 번째 시트: 배치결과 (학생 ID)")
print("  - 두 번째 시트: 명단 (A반, B반 학번 및 이름 입력)")
print("  - 세 번째 시트: 배치결과_이름 (명단 시트의 이름으로 자동 치환)")
if use_row_based_capacity:
    print("  - 네 번째 시트: 학생별배치요약 (전체 기간 정보 및 학생별 배치 정보)")
