# -*- coding: utf-8 -*-
"""
간호학과 78명( A1–A37, B1–B41 )의 세 과목-병원-기간 배치를
모든 제약조건(학생별 3과목 1회, 기간 중복 금지, 병원 중복 금지,
셀 최소 2명, 2학기 셀 의무 사용, Max/Min 용량) 을 만족하도록
정수계획으로 최적화한 뒤 CSV 출력
"""
import csv
from collections import defaultdict, Counter          # ← 0508 오류 원인 보완
from itertools import product
from pathlib import Path

from pulp import (
    LpProblem, LpVariable, LpBinary, lpSum,
    LpStatusOptimal, PULP_CBC_CMD
)

# -------------------- 1. 기본 데이터 --------------------
# 웹 페이지에서 저장한 설정이 있으면 web_config.json에서 불러와 사용하고,
# 없으면 아래의 기본값을 사용합니다.
import json
from pathlib import Path

cfg = None
try:
    p = Path("web_config.json")
    if p.exists():
        with p.open("r", encoding="utf-8") as f:
            cfg = json.load(f)
except Exception:
    cfg = None

if cfg:
    num_periods = int(cfg.get("numPeriods") or cfg.get("oneA", {}).get("numPeriods") or 0)
    periods = [f"T{i}" for i in range(1, num_periods + 1)]

    assignments = cfg.get("assignments", [])
    def get_semester(i: int) -> str:
        for a in assignments:
            if int(a.get("period", 0)) == i:
                return str(a.get("semester") or "")
        return ""
    second_semester = {f"T{i}" for i in range(1, num_periods + 1) if get_semester(i) == "2"}

    numA = int(cfg.get("numA", 0))
    numB = int(cfg.get("numB", 0))
    students_A = [f"A{i}" for i in range(1, numA + 1)]
    students_B = [f"B{i}" for i in range(1, numB + 1)]
    students = students_A + students_B

    def rule_for_period(i: int):
        cls = ""
        for a in assignments:
            if int(a.get("period", 0)) == i:
                cls = (a.get("class") or "").strip()
                break
        if cls == "A":
            return lambda s: s.startswith("A")
        if cls == "B":
            return lambda s: s.startswith("B")
        return lambda s: True  # A&B 또는 미지정 → 모두 가능

    period_ok = {f"T{i}": rule_for_period(i) for i in range(1, num_periods + 1)}

    oneA = cfg.get("oneA", {})
    max_rows = oneA.get("max", [])
    min_rows = oneA.get("min", [])

    def rows_to_csv(rows: list) -> str:
        lines = []
        for r in rows:
            course = r.get("subject", "")
            hosp = r.get("hospital", "")
            values = r.get("values", [])
            # pad/truncate to num_periods
            values = (values + [0] * num_periods)[:num_periods]
            line = ",".join([course, hosp] + [str(int(v)) for v in values])
            lines.append(line)
        return "\n".join(lines)

    RAW_MAX = rows_to_csv(max_rows)
    RAW_MIN = rows_to_csv(min_rows)
else:
    periods = ["T1", "T2", "T3", "T4", "T5", "T6", "T7"]
    second_semester = {"T4", "T5", "T6", "T7"}             # 2학기

    students_A = [f"A{i}" for i in range(1, 38)]
    students_B = [f"B{i}" for i in range(1, 42)]
    students = students_A + students_B

    # (A·B 출석 가능 여부)
    period_ok = {
        "T1": lambda s: s.startswith("B"),
        "T2": lambda s: s.startswith("A"),
        "T3": lambda s: True,
        "T4": lambda s: s.startswith("A"),
        "T5": lambda s: s.startswith("A"),
        "T6": lambda s: s.startswith("B"),
        "T7": lambda s: s.startswith("B"),
    }

    # -------- 1-A. 최대·최소 용량 테이블(문자열 그대로 → 파싱) -------------
    RAW_MAX = """
과목1,병원1,4,4,4,0,0,0,0
과목1,병원1,4,4,4,0,0,0,0
과목1,병원1,4,4,4,0,0,0,0
과목1,병원1,4,4,4,0,0,0,0
과목1,병원2,0,0,0,2,2,2,2
과목1,병원2,0,0,0,2,2,2,2
과목1,병원2,0,0,0,2,2,2,2
과목1,병원3,4,4,0,0,0,0,0
과목1,병원3,4,4,0,0,0,0,0
과목1,병원3,4,4,0,0,0,0,0
과목1,병원3,4,4,0,0,0,0,0
과목2,병원4,4,4,4,0,0,0,0
과목2,병원3,0,4,0,4,0,0,0
과목2,병원6,5,0,0,5,5,5,0
과목2,병원7,0,0,0,4,4,4,4
과목2,병원3,0,4,0,0,0,0,0
과목2,병원8,0,0,0,4,4,4,4
과목2,병원5,8,8,0,6,0,4,0
과목3,병원9,4,4,4,0,0,0,0
과목3,병원3,4,4,4,2,2,2,0
과목3,병원10,4,4,0,4,4,4,4
과목3,병원11,2,2,2,2,2,2,2
과목3,병원11,4,4,4,4,4,4,4
과목3,병원1,4,4,4,0,0,0,0
""".strip()

    RAW_MIN = """
과목1,병원1,0,0,0,0,0,0,0
과목1,병원1,0,0,0,0,0,0,0
과목1,병원1,0,0,0,0,0,0,0
과목1,병원1,0,0,0,0,0,0,0
과목1,병원2,0,0,0,2,2,2,2
과목1,병원2,0,0,0,2,2,2,2
과목1,병원2,0,0,0,2,2,2,2
과목1,병원3,0,0,0,0,0,0,0
과목1,병원3,0,0,0,0,0,0,0
과목1,병원3,0,0,0,0,0,0,0
과목1,병원3,0,0,0,0,0,0,0
과목2,병원4,0,0,0,0,0,0,0
과목2,병원3,0,0,0,4,0,0,0
과목2,병원6,0,0,0,4,4,4,0
과목2,병원7,0,0,0,3,3,3,3
과목2,병원3,0,0,0,0,0,0,0
과목2,병원8,0,0,0,3,3,3,3
과목2,병원5,0,0,0,3,0,3,0
과목3,병원9,0,0,0,0,0,0,0
과목3,병원3,0,0,0,2,2,2,0
과목3,병원10,0,0,0,4,4,4,4
과목3,병원11,0,0,0,2,2,2,2
과목3,병원11,0,0,0,3,3,3,2
과목3,병원1,0,0,0,0,0,0,0
""".strip()

def parse_rows(raw: str):
    rows = []
    for idx, line in enumerate(raw.splitlines()):
        course, hosp, *nums = line.strip().split(",")
        rows.append({
            "id": f"R{idx:02d}",          # 고유 ID
            "course": course,
            "hospital": hosp,
            "capacity": {p: int(x) for p, x in zip(periods, nums)}
        })
    return rows

rows_max = parse_rows(RAW_MAX)
rows_min = parse_rows(RAW_MIN)          # 같은 순서 전제

# ----- 1-B. 셀별 최저 인원 계산(규칙 4·5 반영) -----
for r_max, r_min in zip(rows_max, rows_min):
    for p in periods:
        rule5 = 2 if (p in second_semester and r_max["capacity"][p] > 0) else 0
        r_min["capacity"][p] = max(r_min["capacity"][p], rule5)

# ---------------- 2. MILP 모델 ----------------------
model = LpProblem("Nursing-Student-Scheduling")

# 변수: assign[s, row_id, period]  ∈ {0,1}
assign = {}
for s, r in product(students, rows_max):
    for p in periods:
        if r["capacity"][p] == 0:                # 배치 불가
            continue
        if not period_ok[p](s):                  # 학생 반·기간 호환 안 됨
            continue
        key = (s, r["id"], p)
        assign[key] = LpVariable(f"x_{s}_{r['id']}_{p}", 0, 1, LpBinary)

# 변수: y_(row,period) – 셀 사용 여부 (0: 미사용, 1: 사용)
y = {}
for r in rows_max:
    for p in periods:
        if r["capacity"][p] == 0:
            continue
        y[(r['id'], p)] = LpVariable(f"y_{r['id']}_{p}", 0, 1, LpBinary)

# 2-A. 학생·과목별 1회 수강
# 실제 입력된 모든 과목 추출
all_courses = sorted(set(r['course'] for r in rows_max))
for s in students:
    for course in all_courses:
        model += (
            lpSum(assign[k] for k in assign
                  if k[0] == s and rows_max[[r['id'] for r in rows_max].index(k[1])]['course'] == course) == 1
        )

# 2-B. 학생·기간 중복 금지
for s in students:
    for p in periods:
        model += lpSum(assign[k] for k in assign if k[0] == s and k[2] == p) <= 1

# 2-C. 학생·병원 중복 금지
for s in students:
    for hosp in {r['hospital'] for r in rows_max}:
        model += lpSum(assign[k] for k in assign
                       if k[0] == s and rows_max[[r['id'] for r in rows_max].index(k[1])]['hospital'] == hosp) <= 1

# 2-D. 셀 용량(최대·최소 및 '0 또는 ≥2')
for r_max, r_min in zip(rows_max, rows_min):
    rid = r_max['id']
    for p in periods:
        cap_max = r_max["capacity"][p]
        if cap_max == 0:
            continue
        cap_min = r_min["capacity"][p]
        # assign 합산
        cell_vars = [assign[k] for k in assign if k[1] == rid and k[2] == p]
        model += lpSum(cell_vars) <= cap_max
        model += lpSum(cell_vars) >= cap_min
        # 0 또는 ≥2
        model += lpSum(cell_vars) >= 2 * y[(rid, p)]
        model += lpSum(cell_vars) <= cap_max * y[(rid, p)]

# ---------------- 3. 풀기 ---------------------------
solver = PULP_CBC_CMD(msg=1, timeLimit=120)
status = model.solve(solver)

if status != LpStatusOptimal:
    raise RuntimeError("배치를 찾지 못했습니다. 제약을 완화하거나 용량을 재검토하세요.")

# ---------------- 4. 결과 CSV 작성 -------------------
# 행 순서 유지
header = ["과목명", "병원명"] + periods
csv_rows = []

# 행별-기간별 학생 이름 모으기
who_in = defaultdict(lambda: defaultdict(list))  # who_in[row_id][period] = [학생…]

for (s, rid, p), var in assign.items():
    if var.value() == 1:
        who_in[rid][p].append(s)

for r in rows_max:
    rid = r['id']
    line = [r['course'], r['hospital']]
    for p in periods:
        if r['capacity'][p] == 0:
            line.append("")           # 배치 불가 칸
        else:
            stu_list = " ".join(sorted(who_in[rid][p]))
            line.append(stu_list)
    csv_rows.append(line)

out_path = Path("배치결과.csv")
with out_path.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(csv_rows)

# ---------------- 5. 엑셀 파일 생성 (명단.csv 없이 동작) ----------------
import pandas as pd
from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

input_csv_path = Path("배치결과.csv")
final_xlsx_path = Path("배치결과_최종.xlsx")

if not input_csv_path.exists():
    raise FileNotFoundError("배치결과.csv 파일이 존재하지 않습니다. 4단계 실행 후 다시 시도하세요.")

# 5-1. 배치결과.csv 읽기
with input_csv_path.open("r", encoding="utf-8-sig", newline="") as f:
    reader = csv.reader(f)
    rows = list(reader)

if not rows:
    raise ValueError("배치결과.csv가 비어 있습니다.")

header = rows[0]
body = rows[1:]

# 학생 수 계산 (배치결과에서 최대 번호 추출 또는 기존 변수 사용)
try:
    # Cell 0에서 정의된 변수 사용 시도
    max_A = len(students_A) if 'students_A' in globals() else 0
    max_B = len(students_B) if 'students_B' in globals() else 0
except:
    max_A, max_B = 0, 0

# 배치결과에서 실제 사용된 최대 번호 추출
if max_A == 0 or max_B == 0:
    max_A_num, max_B_num = 0, 0
    for row in body:
        for cell in row[2:]:  # T1, T2, ... 열만 확인
            if cell.strip():
                tokens = cell.strip().split()
                for token in tokens:
                    match = re.match(r'([AB])(\d+)', token)
                    if match:
                        class_letter, num = match.groups()
                        num = int(num)
                        if class_letter == 'A':
                            max_A_num = max(max_A_num, num)
                        else:
                            max_B_num = max(max_B_num, num)
    max_A = max_A_num if max_A == 0 else max_A
    max_B = max_B_num if max_B == 0 else max_B

# 5-2. 엑셀 파일 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 시트 1: 배치결과 (ID 형태)
ws1 = wb.create_sheet("배치결과")
ws1.append(header)
for row in body:
    ws1.append(row)

# 시트 2: 명단 (A반, B반 헤더만)
ws2 = wb.create_sheet("명단")
ws2.append(["A반", "B반"])

# A반, B반 학생 수만큼 빈 행 추가
max_students = max(max_A, max_B)
for i in range(max_students):
    ws2.append(["", ""])

# 시트 3: 배치결과_이름 (엑셀 함수 사용)
ws3 = wb.create_sheet("배치결과_이름")
ws3.append(header)

# 각 셀에 엑셀 함수 입력
for row_idx, row in enumerate(body, start=2):  # 2부터 시작 (헤더 다음)
    new_row = [row[0], row[1]]  # 과목명, 병원명은 그대로
    
    for col_idx in range(2, len(row)):  # T1, T2, ... 열
        cell_value = row[col_idx].strip()
        
        if not cell_value:
            # 빈 셀
            new_row.append("")
        else:
            # 여러 ID가 공백으로 구분된 경우 처리
            tokens = cell_value.split()
            
            # 각 ID를 이름으로 변환하는 함수 생성
            formula_parts = []
            
            for token in tokens:
                # A1, A2, B1, B2 형식 파싱
                match = re.match(r'([AB])(\d+)', token)
                if match:
                    class_letter, num_str = match.groups()
                    num = int(num_str)
                    
                    ref_col = class_letter  # A 또는 B
                    ref_row = num + 1  # 헤더 다음부터 (A1 -> 2행, A2 -> 3행)
                    
                    # IF(ISBLANK(명단!A2), "", 명단!A2) 형태
                    ref_cell = f"명단!${ref_col}${ref_row}"
                    formula_parts.append(f'IF(ISBLANK({ref_cell}), "", {ref_cell})')
                else:
                    # 알 수 없는 형식은 그대로 (따옴표 처리)
                    formula_parts.append(f'"{token}"')
            
            if len(formula_parts) == 0:
                formula = '""'
            elif len(formula_parts) == 1:
                # 단일 ID
                formula = formula_parts[0]
            else:
                # 여러 ID를 공백으로 연결
                formula = '&" "&'.join(formula_parts)
            
            new_row.append(f"={formula}")
    
    ws3.append(new_row)

# 5-3. 엑셀 파일 저장
wb.save(final_xlsx_path)
print(f"완료: {final_xlsx_path.resolve()}")
print("  - 시트1 '배치결과': 배치 결과 (ID 형태)")
print("  - 시트2 '명단': A반, B반 헤더만 (이름 입력용)")
print("  - 시트3 '배치결과_이름': 엑셀 함수로 이름 자동 표시")
